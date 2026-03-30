import os
import pandas as pd
from openpyxl import load_workbook


def append_to_excel(df: pd.DataFrame, file_path: str, sheet_name: str = "Rechnungen"):
    if not os.path.exists(file_path):
        df.to_excel(file_path, index=False, sheet_name=sheet_name)
        return

    book = load_workbook(file_path)

    if sheet_name not in book.sheetnames:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        return

    ws = book[sheet_name]

    for row in df.itertuples(index=False, name=None):
        ws.append(row)

    book.save(file_path)